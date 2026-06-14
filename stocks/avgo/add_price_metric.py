#!/usr/bin/env python3
"""Add Stock Price as a FIN_METRICS entry using avgo-10-year-history.csv."""
import csv, json, re, datetime
from pathlib import Path
import openpyxl

CSV  = Path(__file__).parent / "avgo-10-year-history.csv"
XLSX = Path(__file__).parent / "avgo-wise-q2fy26.xlsx"
HTML = Path(__file__).parent / "avgo-thesis.html"

# ── Load CSV → date:price dict ───────────────────────────────────────────────
prices = {}
with open(CSV, newline="", encoding="utf-8-sig") as f:
    for row in csv.reader(f):
        if len(row) >= 2:
            try:
                dt = datetime.date.fromisoformat(row[0].strip())
                prices[dt] = round(float(row[1].strip()), 2)
            except:
                pass

print(f"Loaded {len(prices)} daily prices: {min(prices)} → {max(prices)}")

# ── Lookup: closest trading day on or before target date ────────────────────
def price_on(target_date):
    if isinstance(target_date, datetime.datetime):
        target_date = target_date.date()
    # Walk back up to 5 days to find a trading day
    for offset in range(6):
        d = target_date - datetime.timedelta(days=offset)
        if d in prices:
            return prices[d]
    return None

# ── Get fiscal period dates from xlsx ───────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)

ws_q  = wb["AVGO - Key Metrics Q"]
ws_fy = wb["AVGO - Key Metrics FY"]

rows_q  = list(ws_q.iter_rows(values_only=True))
rows_fy = list(ws_fy.iter_rows(values_only=True))

# Row 5 (index 4) = Date row
q_dates_raw  = [d for d in rows_q[4][1:] if d][:20]   # 20 most recent quarters
fy_dates_raw = [d for d in rows_fy[4][1:] if d][:10]  # 10 most recent FY

# Both lists are newest-first; reverse to get oldest→newest
q_dates  = list(reversed(q_dates_raw))   # Q3FY21 → Q2FY26
fy_dates = list(reversed(fy_dates_raw))  # FY16 → FY25

print("\nQuarterly dates (Q3FY21→Q2FY26):")
for d in q_dates:
    p = price_on(d)
    print(f"  {d.strftime('%Y-%m-%d') if hasattr(d,'strftime') else d}  →  ${p}")

print("\nAnnual dates (FY16→FY25):")
for d in fy_dates:
    p = price_on(d)
    print(f"  {d.strftime('%Y-%m-%d') if hasattr(d,'strftime') else d}  →  ${p}")

# ── Build metric ─────────────────────────────────────────────────────────────
q_prices  = [price_on(d) for d in q_dates]
fy_prices = [price_on(d) for d in fy_dates]

metric = {
    "name": "Stock Price",
    "group": "Valuation",
    "fmt": "$",
    "desc": "Closing price on fiscal period end date",
    "annual":    fy_prices,
    "quarterly": q_prices
}

nulls_a = sum(1 for v in fy_prices if v is None)
nulls_q = sum(1 for v in q_prices  if v is None)
print(f"\nMetric: annual={fy_prices} ({nulls_a} null)")
print(f"        quarterly last 5={q_prices[-5:]} ({nulls_q} null)")

# ── Inject before closing ]; of FIN_METRICS ─────────────────────────────────
src = HTML.read_text(encoding="utf-8")

fin_start = src.find("const FIN_METRICS = [")
depth = 0
close_idx = None
for i, ch in enumerate(src[fin_start:], start=fin_start):
    if ch == '[': depth += 1
    elif ch == ']':
        depth -= 1
        if depth == 0:
            close_idx = i
            break

ann_str = json.dumps(metric["annual"])
qrt_str = json.dumps(metric["quarterly"])
js = f'''  {{
    "name": "{metric['name']}",
    "group": "{metric['group']}",
    "fmt": "{metric['fmt']}",
    "desc": "{metric['desc']}",
    "annual": {ann_str},
    "quarterly": {qrt_str}
  }}'''

new_src = src[:close_idx] + ",\n" + js + "\n" + src[close_idx:]
HTML.write_text(new_src, encoding="utf-8")

# ── Verify ───────────────────────────────────────────────────────────────────
txt = HTML.read_text(encoding="utf-8")
found = '"name": "Stock Price"' in txt
total = txt.count('"name":')
print(f"\n{'OK' if found else 'MISS'}: Stock Price injected")
print(f"Total metrics in FIN_METRICS: {total}")
print(f"File size: {len(txt):,} bytes")
